# IMPORTAÇÕES E BIBLIOTECAS
import os, json, time, re, threading, queue
from datetime import datetime, timedelta
from nicegui import ui, app
import serial
import serial.tools.list_ports
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# CONFIGURAÇÕES E CONSTANTES
ARQ_FUNC = "funcionarios.json"
ARQ_REG  = "registros.json"

BAUDRATE = 9600
TIMEOUT = 1
EVENTOS = ["entrada", "saida_intervalo", "volta_intervalo", "saida"]
MIN_GAP_SECONDS = 60
HEX_RE = re.compile(r'^[0-9A-F]+$')

# VARIAVEIS GLOBAIS
funcionarios = {}
registros = {}
ultimas_batidas = {}
serial_thread = None
serial_stop_flag = threading.Event()
serial_port = None
serial_connected = False
serial_queue = queue.Queue()    # ('ok'|'err'|'log'|'uid_captured', payload)
PORTA_ATUAL = None
last_export_path = None

capture_uid_mode = False
capture_lock = threading.Lock()

# FUNÇÕES AUXILIARES
def carregar_json(path, default):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default

def salvar_json(path, data):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def agora():
    dt = datetime.now()
    return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M")

def proximo_evento(dia_dict):
    for ev in EVENTOS:
        if ev not in dia_dict:
            return ev
    return None

def extrair_uid(linha: str):
    """Retorna UID válido (HEX, tamanho par entre 8 e 20) ou None."""
    if not linha:
        return None
    s = linha.strip().upper()
    if s.startswith("#") or s in {"READY", "OK", "ERR"}:
        return None
    if s.startswith("UID:"):
        s = s.split(":", 1)[1].strip()
    if HEX_RE.match(s) and (8 <= len(s) <= 20) and (len(s) % 2 == 0):
        return s
    return None

def registrar_batida(uid):
    """Registra batida e retorna (ok: bool, msg: str, evento_ou_ERR: str)."""
    uid = uid.strip().upper()
    if not uid:
        return False, "UID vazio", "ERR"

    t = time.time()
    if uid in ultimas_batidas and (t - ultimas_batidas[uid]) < MIN_GAP_SECONDS:
        return False, f"Toque repetido em < {MIN_GAP_SECONDS}s", "ERR"
    ultimas_batidas[uid] = t

    if uid not in funcionarios:
        return False, "UID não cadastrado", "ERR"

    data_str, hora_str = agora()
    if uid not in registros:
        registros[uid] = {}
    if data_str not in registros[uid]:
        registros[uid][data_str] = {}

    dia = registros[uid][data_str]
    ev = proximo_evento(dia)
    if ev is None:
        return False, "Dia já completo", "ERR"

    dia[ev] = hora_str
    salvar_json(ARQ_REG, registros)
    return True, f"{funcionarios[uid]}: {ev.replace('_',' ')} às {hora_str} ({data_str})", ev

def listar_portas():
    return [p.device for p in serial.tools.list_ports.comports()]

# HELPERS PARA EXPORTAÇÃO EXCEL
def _parse_hhmm(s):
    try:
        return datetime.strptime(s, "%H:%M").time()
    except Exception:
        return None

def calcular_horas_dia_excel(dia: dict) -> float:
    """
    Calcula horas do dia em FRAÇÃO DE DIA (padrão Excel):
    (saida - entrada) - (volta_intervalo - saida_intervalo)
    """
    e = _parse_hhmm(dia.get('entrada', ''))
    s = _parse_hhmm(dia.get('saida', ''))
    if not (e and s):
        return 0.0
    si = _parse_hhmm(dia.get('saida_intervalo', ''))
    vi = _parse_hhmm(dia.get('volta_intervalo', ''))

    base = (datetime.combine(datetime.today(), s) -
            datetime.combine(datetime.today(), e))
    if si and vi:
        base -= (datetime.combine(datetime.today(), vi) -
                 datetime.combine(datetime.today(), si))
    secs = max(0, base.total_seconds())
    return secs / 86400.0  # fração de dia

def exportar_mes_xlsx(ano_mes: str, funcionarios: dict, registros: dict, eventos: list[str]) -> str:
    """
    Gera export/<MM-YYYY>_registros.xlsx
    - Uma aba por funcionário (com dados do mês)
    - Aba 'Resumo' com total de horas por funcionário
    Retorna caminho do arquivo gerado.
    """
    if len(ano_mes) != 7 or ano_mes[4] != "-":
        raise ValueError("Use o formato YYYY-MM (ex.: 2025-11)")

    os.makedirs("export", exist_ok=True)
    nome_arq = f"Registros_{datetime.strptime(ano_mes, '%Y-%m').strftime('%m-%Y')}.xlsx"
    caminho = os.path.join("export", nome_arq)


    wb = Workbook()
    # Estilos
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    totais_func = []
    created_any = False

    for uid, nome in sorted(funcionarios.items(), key=lambda x: x[1].lower()):
        dias_mes = {d: dia for d, dia in (registros.get(uid, {}) or {}).items()
                    if d.startswith(ano_mes)}
        if not dias_mes:
            continue

        created_any = True
        title_base = nome.strip().replace("/", "-").replace("\\", "-").replace(":", "-")
        title = title_base[:31] if title_base else uid[:8]
        ws = wb.create_sheet(title=title)

        cols = ["Data", "Dia", "Entrada", "Saída Intervalo", "Volta Intervalo", "Saída", "Horas"]
        ws.append(cols)
        for col_idx in range(1, len(cols)+1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border_thin

        total_horas = 0.0
        for data_str in sorted(dias_mes.keys()):
            dia = dias_mes[data_str]
            try:
                dt = datetime.strptime(data_str, "%Y-%m-%d")
                dia_semana = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"][dt.weekday()]
            except Exception:
                dia_semana = ""
            
            data_br = datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
            row = [
                data_br,
                dia_semana,
                dia.get("entrada", ""),
                dia.get("saida_intervalo", ""),
                dia.get("volta_intervalo", ""),
                dia.get("saida", ""),
                None,
            ]
            ws.append(row)

            r = ws.max_row
            horas_frac = calcular_horas_dia_excel(dia)
            total_horas += horas_frac
            c = ws.cell(row=r, column=7, value=horas_frac)
            c.number_format = "[h]:mm"
            c.alignment = right
            c.border = border_thin

            for col in range(1, 7):
                cell = ws.cell(row=r, column=col)
                cell.alignment = center
                cell.border = border_thin

        ws.append(["", "", "", "", "TOTAL", "", total_horas])
        r = ws.max_row
        ws.cell(row=r, column=5).font = Font(bold=True)
        tc = ws.cell(row=r, column=7)
        tc.number_format = "[h]:mm"
        tc.font = Font(bold=True)
        tc.alignment = right
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border_thin

        widths = [11, 6, 10, 16, 16, 10, 9]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:G{ws.max_row}"

        totais_func.append((nome, uid, total_horas))

    if "Sheet" in wb.sheetnames and not created_any:
        ws = wb["Sheet"]
        ws.title = "Resumo"
        ws.append(["Sem registros para", ano_mes])
    else:
        if "Sheet" in wb.sheetnames:
            std = wb["Sheet"]
            wb.remove(std)

    ws_r = wb.create_sheet("Resumo", 0)
    ws_r.append(["Funcionário", "UID", "Total Horas"])
    for col_idx in range(1, 4):
        c = ws_r.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border_thin

    for nome, uid, total in sorted(totais_func, key=lambda x: x[0].lower()):
        ws_r.append([nome, uid, total])
        r = ws_r.max_row
        ws_r.cell(row=r, column=3).number_format = "[h]:mm"
        for col in range(1, 4):
            ws_r.cell(row=r, column=col).border = border_thin
            ws_r.cell(row=r, column=col).alignment = center

    ws_r.column_dimensions["A"].width = 36
    ws_r.column_dimensions["B"].width = 20
    ws_r.column_dimensions["C"].width = 12
    ws_r.freeze_panes = "A2"
    ws_r.auto_filter.ref = f"A1:C{ws_r.max_row}"

    wb.save(caminho)
    return caminho

# FUNÇÕES DE THREAD SERIAL
def serial_worker(port_name):
    global serial_connected, serial_port, capture_uid_mode
    try:
        with serial.Serial(port_name, BAUDRATE, timeout=TIMEOUT) as ar:
            serial_port = ar
            serial_connected = True
            serial_queue.put(("log", f"[SERIAL] Conectado em {port_name} @ {BAUDRATE}"))
            while not serial_stop_flag.is_set():
                try:
                    linha = ar.readline().decode("utf-8", errors="ignore").strip()
                    if not linha:
                        continue
                    uid = extrair_uid(linha)
                    if uid is None:
                        continue
                    # modo captura
                    with capture_lock:
                        if capture_uid_mode:
                            try:
                                ar.write(b"OK\n")
                            except Exception as ew:
                                serial_queue.put(("log", f"[WARN] Falha ACK (captura): {ew}"))
                            serial_queue.put(("uid_captured", uid))
                            capture_uid_mode = False
                            continue
                    # modo normal
                    ok, info, evento = registrar_batida(uid)
                    try:
                        ar.write(b"OK\n" if ok else b"ERR\n")
                    except Exception as ew:
                        serial_queue.put(("log", f"[WARN] Falha ao enviar ACK: {ew}"))
                    if ok:
                        serial_queue.put(("ok", f"[OK] {info}"))
                    else:
                        serial_queue.put(("err", f"[ERR] UID {uid}: {info}"))
                except Exception as e:
                    serial_queue.put(("log", f"[WARN] Leitura: {e}"))
                    time.sleep(0.3)
    except Exception as e:
        serial_queue.put(("log", f"[ERRO] Não abriu {port_name}: {e}"))
    finally:
        serial_connected = False
        serial_port = None
        serial_queue.put(("log", "[SERIAL] Desconectado"))

# ===================== Carrega dados na inicialização =====================
funcionarios = carregar_json(ARQ_FUNC, {})
registros = carregar_json(ARQ_REG, {})

# ===================== UI (NiceGUI) =====================
with ui.header().classes(replace='row items-center justify-between'):
    ui.button(icon='menu').props('flat color=white')
    with ui.tabs() as tabs:
        ui.tab('Conexão')
        ui.tab('Cadastro')
        ui.tab('Remover')
        ui.tab('Registros por Funcionário')
        ui.tab('Registros Totais (Dia)')
        ui.tab('Exportar Excel')
    status_label = ui.label('Desconectado').classes('text-red-600')

with ui.footer(value=False) as footer:
    ui.label('Trabalho Microcontroladores • Ponto NFC')

# ---------- Painéis por aba ----------
with ui.tab_panels(tabs, value='Conexão').classes('w-full'):
    # ====== ABA CONEXÃO ======
    with ui.tab_panel('Conexão'):
        with ui.row().classes('w-full items-end gap-4'):
            portas_select = ui.select(options=listar_portas(), label='Porta Serial', with_input=True)\
                              .classes('min-w-[220px]')
            portas_select.value = portas_select.options[0] if portas_select.options else None

            def refresh_ports():
                portas_select.options = listar_portas()
                ui.notify('Portas atualizadas', type='positive')
            ui.button('Atualizar portas', on_click=refresh_ports)

            def conectar():
                global serial_thread, PORTA_ATUAL
                if serial_connected:
                    ui.notify('Já conectado', type='warning'); return
                if not portas_select.value:
                    ui.notify('Selecione uma porta', type='warning'); return
                PORTA_ATUAL = portas_select.value
                serial_stop_flag.clear()
                serial_thread = threading.Thread(target=serial_worker, args=(PORTA_ATUAL,), daemon=True)
                serial_thread.start()
                ui.notify(f'Conectando em {PORTA_ATUAL}...', type='info')

            def desconectar():
                if not serial_connected:
                    ui.notify('Já desconectado', type='warning'); return
                serial_stop_flag.set()
                ui.notify('Desconectando...', type='info')

            ui.button('Conectar', on_click=conectar, color='green')
            ui.button('Desconectar', on_click=desconectar, color='red')

    # ====== ABA CADASTRO ======
    with ui.tab_panel('Cadastro'):
        ui.label('Cadastrar funcionário').classes('text-lg font-medium')
        with ui.row().classes('items-end gap-3'):
            nome_in = ui.input('Nome').classes('min-w-[260px]')
            uid_in  = ui.input('UID (hex)').classes('min-w-[260px]')

            def capturar_uid():
                global capture_uid_mode
                if not serial_connected:
                    ui.notify('Conecte à serial para capturar UID', type='warning'); return
                with capture_lock:
                    capture_uid_mode = True
                ui.notify('Aproxime o cartão para capturar UID...', type='info')

            ui.button('Capturar próximo UID', on_click=capturar_uid, icon='fingerprint')

        def salvar_funcionario():
            nome = (nome_in.value or '').strip()
            uid  = (uid_in.value or '').strip().upper()
            if not nome or not uid:
                ui.notify('Preencha nome e UID', type='warning'); return
            if not HEX_RE.match(uid):
                ui.notify('UID inválido (use somente HEX)', type='warning'); return
            if uid in funcionarios:
                ui.notify('UID já cadastrado', type='warning'); return
            funcionarios[uid] = nome
            salvar_json(ARQ_FUNC, funcionarios)
            ui.notify(f'Cadastrado: {nome} ({uid})', type='positive')
            atualizar_remover_ui()
            atualizar_tabela_batidas_por_func()
            nome_in.value = ''
            uid_in.value = ''

        ui.button('Salvar', on_click=salvar_funcionario, color='green')

    # ====== ABA REMOVER ======
    with ui.tab_panel('Remover'):
        ui.label('Remover funcionário').classes('text-lg font-medium')

        def _options_por_nome():
            contagem = {}
            for uid, nome in funcionarios.items():
                contagem[nome] = contagem.get(nome, 0) + 1
            opts = {}
            for uid, nome in sorted(funcionarios.items(), key=lambda x: x[1].lower()):
                label = nome if contagem[nome] == 1 else f'{nome} ({uid[:6]}...)'
                opts[uid] = label
            return opts

        sel_nome = ui.select(options=_options_por_nome(), label='Selecione pelo nome').classes('min-w-[420px]')
        apagar_chk = ui.checkbox('Apagar também os registros', value=False)

        def atualizar_remover_ui():
            sel_nome.options = _options_por_nome()
            sel_nome.value = None
            sel_nome.update()

        def remover_agora():
            uid = sel_nome.value
            if not uid:
                ui.notify('Selecione um funcionário', type='warning'); return
            nome = funcionarios.get(uid)
            if not nome:
                ui.notify('Funcionário não encontrado', type='warning'); return
            funcionarios.pop(uid, None)
            salvar_json(ARQ_FUNC, funcionarios)
            if apagar_chk.value:
                registros.pop(uid, None)
                salvar_json(ARQ_REG, registros)
            ui.notify(f'Funcionário "{nome}" removido do sistema.', type='positive')
            atualizar_remover_ui()
            try: atualizar_tabela_batidas_por_func()
            except: pass
            try: atualizar_lobby_table()
            except: pass

        ui.button('Remover funcionário', color='red', on_click=remover_agora)

# ====== ABA REGISTROS ======
    with ui.tab_panel('Registros por Funcionário'):
        ui.label('Registros por funcionário').classes('text-lg font-medium')

        # --- helpers de datas ---
        def coletar_datas_disponiveis():
            """Retorna (options_dict, default_iso): options = {ISO: 'DD/MM/AAAA'}, ordenadas por mais recente."""
            datas = set()
            for _uid, dias in registros.items():
                datas.update(dias.keys())
            if not datas:
                hoje_iso = datetime.now().strftime("%Y-%m-%d")
                return {hoje_iso: datetime.strptime(hoje_iso, "%Y-%m-%d").strftime("%d/%m/%Y")}, hoje_iso

            # ordena ISO desc (mais recente primeiro)
            ordenadas = sorted(datas, reverse=True)
            options = {iso: datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y") for iso in ordenadas}
            return options, ordenadas[0]  # default = mais recente

        datas_options, default_iso = coletar_datas_disponiveis()

        # select de datas (value = ISO; label = BR)
        datas_select = ui.select(
            options=datas_options,
            value=default_iso,
            label='Data'
        ).classes('min-w-[210px]')

        batidas_container = ui.column().classes('w-full')

        def desenhar_batidas_por_func():
            batidas_container.clear()
            data_iso = datas_select.value or datetime.now().strftime("%Y-%m-%d")
            data_br = datetime.strptime(data_iso, "%Y-%m-%d").strftime("%d/%m/%Y")

            for uid, nome in sorted(funcionarios.items(), key=lambda x: x[1].lower()):
                dia = registros.get(uid, {}).get(data_iso, {})
                with batidas_container:
                    with ui.expansion(f'{nome} ({uid}) - {data_br}', value=False).classes('w-full'):
                        with ui.card().classes('w-full'):
                            rows = [{
                                'entrada': dia.get('entrada', ''),
                                'saida_intervalo': dia.get('saida_intervalo', ''),
                                'volta_intervalo': dia.get('volta_intervalo', ''),
                                'saida': dia.get('saida', ''),
                            }]
                            ui.table(
                                columns=[
                                    {'name': 'entrada', 'label': 'Entrada', 'field': 'entrada'},
                                    {'name': 'saida_intervalo', 'label': 'Saída Intervalo', 'field': 'saida_intervalo'},
                                    {'name': 'volta_intervalo', 'label': 'Volta Intervalo', 'field': 'volta_intervalo'},
                                    {'name': 'saida', 'label': 'Saída', 'field': 'saida'},
                                ],
                                rows=rows,
                            ).classes('w-full')

        def atualizar_datas_select():
            """Recarrega as datas disponíveis e mantém a seleção quando possível."""
            old_val = datas_select.value
            options, default_iso_local = coletar_datas_disponiveis()
            datas_select.options = options
            # tenta manter a data selecionada; se não existir mais, usa a mais recente
            datas_select.value = old_val if old_val in options else default_iso_local
            datas_select.update()

        def atualizar_tabela_batidas_por_func():
            desenhar_batidas_por_func()

        # quando trocarem a data, redesenha
        datas_select.on_value_change(lambda _: desenhar_batidas_por_func())

        # primeiro desenho
        desenhar_batidas_por_func()

    # ====== ABA REGISTROS TOTAIS ======
    with ui.tab_panel('Registros Totais (Dia)'):
        ui.label(f'Registros do dia ({datetime.now().strftime("%d/%m/%Y")})').classes('text-lg font-medium')
        with ui.row().classes('w-full'):
            lobby_table = ui.table(
                columns=[
                    {'name': 'hora', 'label': 'Hora', 'field': 'hora'},
                    {'name': 'nome', 'label': 'Nome', 'field': 'nome'},
                    {'name': 'evento', 'label': 'Evento', 'field': 'evento'},
                    {'name': 'uid', 'label': 'UID', 'field': 'uid'},
                ],
                rows=[],
            ).classes('w-1/2')

        def atualizar_lobby_table():
            hoje = datetime.now().strftime("%Y-%m-%d")
            items = []
            for uid, dias in registros.items():
                nome = funcionarios.get(uid, uid)
                dia = dias.get(hoje, {})
                for ev in EVENTOS:
                    if ev in dia:
                        items.append({'hora': dia[ev], 'nome': nome, 'evento': ev.replace('_', ' '), 'uid': uid})
            items.sort(key=lambda r: r['hora'])
            lobby_table.rows = items
            lobby_table.update()

        ui.button('Atualizar', on_click=atualizar_lobby_table)
        atualizar_lobby_table()

    # ====== ABA EXPORTAR EXCEL ======


    with ui.tab_panel('Exportar Excel'):
        ui.label('Exportar registros para Excel (.xlsx)').classes('text-lg font-medium')
        def coletar_meses_disponiveis():
            """Retorna (options_dict, default_iso):
            options_dict = { 'YYYY-MM': 'MM/YYYY' }, ordenado do mais recente para o mais antigo."""
            meses = set()
            for _uid, dias in registros.items():
                for data_iso in dias.keys():          # data_iso = 'YYYY-MM-DD'
                    if len(data_iso) >= 7:
                        meses.add(data_iso[:7])       # 'YYYY-MM'
            if not meses:
                atual_iso = datetime.now().strftime("%Y-%m")
                return {atual_iso: datetime.strptime(atual_iso, "%Y-%m").strftime("%m/%Y")}, atual_iso

            ordenados = sorted(meses, reverse=True)   # mais recente primeiro
            options = {m: datetime.strptime(m, "%Y-%m").strftime("%m/%Y") for m in ordenados}
            return options, ordenados[0]

        mes_options, mes_default = coletar_meses_disponiveis()
        mes_select = ui.select(options=mes_options, value=mes_default, label='Mês (MM/YYYY)').classes('min-w-[200px]')
        abrir_btn = ui.button('Abrir no Excel', icon='folder_open', on_click=lambda: abrir_no_excel())
        abrir_btn.disable()  # começa desativado
        
        def abrir_no_excel():
            import os, subprocess, sys
            global last_export_path
            if not last_export_path or not os.path.exists(last_export_path):
                ui.notify('Nenhum arquivo exportado ainda ou arquivo não encontrado.', type='warning')
                return
            try:
                if sys.platform.startswith("win"):
                    os.startfile(last_export_path)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", last_export_path])
                else:
                    subprocess.Popen(["xdg-open", last_export_path])
                ui.notify(f'Abrindo: {os.path.basename(last_export_path)}', type='positive')
            except Exception as e:
                ui.notify(f'Erro ao abrir: {e}', type='negative')

        def exportar_xlsx_ui():
            import os, sys, subprocess
            global last_export_path
            mes_iso = (mes_select.value or '').strip()   # já é 'YYYY-MM'
            try:
                caminho = exportar_mes_xlsx(mes_iso, funcionarios, registros, EVENTOS)
                ui.notify(f'Arquivo gerado: {caminho}', type='positive')

                # dispara o download pelo navegador (mantém o comportamento atual)
                ui.download(caminho)

                # guarda caminho absoluto para o botão "Abrir no Excel"
                last_export_path = os.path.abspath(caminho)
                abrir_btn.enable()  # habilita o botão agora que existe um arquivo

            except Exception as e:
                ui.notify(f'Erro: {e}', type='negative')

        ui.button('Exportar mês (xlsx)', on_click=exportar_xlsx_ui, color='primary')
        ui.label('Gera 1 arquivo por mês, com 1 aba por funcionário + aba Resumo. "Horas" no formato [h]:mm.')

# ====== Downloads (estáticos) ======
app.add_static_files('/data', '.')   # /data/funcionarios.json, /data/registros.json e /data/export/...

# ===================== Timers e Handlers =====================
def push_log(texto, tipo="info"):
    if tipo == "ok":
        ui.notify(texto, type='positive', position='top-right')
    elif tipo == "err":
        ui.notify(texto, type='negative', position='top-right')

def ui_tick():
    # status destacado
    if serial_connected:
        status_label.text = f'CONECTADO ({PORTA_ATUAL})'
        status_label.classes(replace='text-white bg-green-600 px-3 py-1 rounded font-bold shadow')
    else:
        status_label.text = 'DESCONECTADO'
        status_label.classes(replace='text-white bg-red-600 px-3 py-1 rounded font-bold shadow pulse')

    # helper para atualizar telas/listas após um evento de leitura
    def _refresh_views():
        # atualiza listas/tabelas das outras abas
        try:
            atualizar_tabela_batidas_por_func()
        except:
            pass
        try:
            atualizar_lobby_table()
        except:
            pass
        # (opcional) se você criou a lista de datas na aba "Registros por Funcionário"
        try:
            atualizar_datas_select()
        except:
            pass
        # atualiza lista de meses na aba "Exportar Excel"
        try:
            new_opts, new_default = coletar_meses_disponiveis()
            mes_select.options = new_opts
            if mes_select.value not in new_opts:
                mes_select.value = new_default
            mes_select.update()
        except Exception as e:
            print(f"[WARN] Falha ao atualizar lista de meses: {e}")

    try:
        while True:
            kind, payload = serial_queue.get_nowait()

            if kind == "ok":
                push_log(payload, "ok")
                _refresh_views()

            elif kind == "err":
                push_log(payload, "err")
                _refresh_views()  # <<< também atualiza no erro

            elif kind == "log":
                push_log(payload, "info")

            elif kind == "uid_captured":
                uid_in.value = payload
                uid_in.update()
                ui.notify(f'UID capturado: {payload}', type='positive')

    except queue.Empty:
        pass

ui.timer(0.2, ui_tick)
ui.run(title='Ponto NFC', reload=False)
